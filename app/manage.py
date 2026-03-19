from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from sqlalchemy import select

from app.db import SessionLocal
from app.models import Company, CrawlSource, JobSource, User
from app.security import hash_password


def cmd_create_user(args: argparse.Namespace) -> None:
    db = SessionLocal()
    try:
        exists = db.execute(select(User).where(User.username == args.username)).scalar_one_or_none()
        if exists:
            raise SystemExit("username already exists")
        u = User(
            username=args.username,
            password_hash=hash_password(args.password),
            is_admin=bool(args.admin),
        )
        db.add(u)
        db.commit()
        print(f"created user {u.username} admin={u.is_admin}")
    finally:
        db.close()


def cmd_set_password(args: argparse.Namespace) -> None:
    db = SessionLocal()
    try:
        u = db.execute(select(User).where(User.username == args.username)).scalar_one_or_none()
        if not u:
            raise SystemExit("user not found")
        u.password_hash = hash_password(args.password)
        db.add(u)
        db.commit()
        print(f"updated password for {u.username}")
    finally:
        db.close()


def cmd_ensure_user(args: argparse.Namespace) -> None:
    """Create user if missing; otherwise reset password and (optionally) admin flag.

    This is intentionally convenient for local/dev use.
    """

    db = SessionLocal()
    try:
        u = db.execute(select(User).where(User.username == args.username)).scalar_one_or_none()
        if not u:
            u = User(
                username=args.username,
                password_hash=hash_password(args.password),
                is_admin=bool(args.admin),
            )
            db.add(u)
            db.commit()
            print(f"created user {u.username} admin={u.is_admin}")
            return

        u.password_hash = hash_password(args.password)
        if args.admin:
            u.is_admin = True
        db.add(u)
        db.commit()
        print(f"ensured user {u.username} admin={u.is_admin} (password reset)")
    finally:
        db.close()


def _load_json(path: str) -> object:
    p = Path(path)
    raw = p.read_text(encoding="utf-8").strip()
    return json.loads(raw) if raw else []


def _read_text_best_effort(path: str) -> str:
    """Read a local text file with best-effort decoding.

    This repo is often used on Windows + PowerShell where console codepages can
    make it hard to embed Chinese literals. We keep file decoding resilient.
    """

    b = Path(path).read_bytes()
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    # Fallback: keep something rather than crash.
    return b.decode("utf-8", errors="ignore")


# ASCII-only URL token (stops before any Chinese chars when records are concatenated).
# Keep '-' first in the character class to avoid "bad character range".
_ASCII_URL_RE = r"https?://[-A-Za-z0-9._~:/?#\[\]@!$&'()*+,;=%]+"


def _parse_company_tsv_text(text: str) -> list[dict]:
    """Parse tab-separated company list text.

    Expected columns (TSV):
    - 行业大类
    - 细分赛道
    - 企业/机构名称
    - 重点招聘方向
    - 官方招聘入口 (URL)

    The file may contain malformed newlines; we match records by tab structure.
    """

    t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    def _clean_major(s: str) -> str:
        # Normalize "一、互联网与科技" -> "互联网与科技"
        v = (s or "").strip()
        v = re.sub(r"^[一二三四五六七八九十]+[、.]\s*", "", v)
        v = re.sub(r"^\d+[、.]\s*", "", v)
        return v.strip()

    def _infer_company_type(major: str) -> str | None:
        industry_major = major or ""
        if "央企" in industry_major or "国企" in industry_major:
            return "央企/国企"
        if "科研" in industry_major or "研究" in industry_major:
            return "科研院所"
        if "互联网" in industry_major or "科技" in industry_major:
            return "科技公司"
        if "全球" in industry_major or "跨国" in industry_major:
            return "外企"
        return None

    def _row(industry_major: str, track: str | None, name: str, focus: str | None, url: str) -> dict | None:
        nm = (name or "").strip()
        u = (url or "").strip()
        if not nm or not u:
            return None
        # Skip header-ish rows.
        if ("企业" in nm and "名称" in nm) or ("官方" in nm and "入口" in nm):
            return None

        # Normalize "百度 (Baidu)" -> "百度" to avoid duplicate companies when the list mixes CN+EN aliases.
        nm = re.sub(r"\s+", " ", nm).strip()
        for sep in ["（", "("]:
            if sep in nm:
                head = nm.split(sep, 1)[0].strip()
                if head and len(head) >= 2:
                    nm = head
                break

        major = _clean_major(industry_major)
        tr = (track or "").strip()
        company_type = _infer_company_type(major)

        industry = major
        if tr and tr not in major:
            industry = f"{major}/{tr}"

        return {
            "name": nm,
            "industry": industry[:120] if industry else None,
            "company_type": company_type,
            "focus_directions": (focus[:200] if focus else None),
            "recruitment_url": u,
            "website": None,
            "_industry_major": major,
        }

    out: list[dict] = []

    # 5 columns: major, track, name, focus, url
    pat5 = re.compile(rf"([^\t\n]+)\t([^\t\n]+)\t([^\t\n]+)\t([^\t\n]+)\t({_ASCII_URL_RE})")
    for m in pat5.finditer(t):
        rec = _row(m.group(1) or "", m.group(2) or "", m.group(3) or "", m.group(4) or "", m.group(5) or "")
        if rec:
            out.append(rec)

    # 4 columns: major, track, name, url (some lists omit focus column)
    pat4 = re.compile(rf"([^\t\n]+)\t([^\t\n]+)\t([^\t\n]+)\t({_ASCII_URL_RE})")
    for m in pat4.finditer(t):
        # Avoid re-adding rows already captured by the 5-col matcher.
        url = (m.group(4) or "").strip()
        if any(url == (it.get("recruitment_url") or "") for it in out):
            continue
        rec = _row(m.group(1) or "", m.group(2) or "", m.group(3) or "", None, url)
        if rec:
            out.append(rec)

    # 3 columns: major, name, url (rare)
    pat3 = re.compile(rf"([^\t\n]+)\t([^\t\n]+)\t({_ASCII_URL_RE})")
    for m in pat3.finditer(t):
        url = (m.group(3) or "").strip()
        if any(url == (it.get("recruitment_url") or "") for it in out):
            continue
        rec = _row(m.group(1) or "", None, m.group(2) or "", None, url)
        if rec:
            out.append(rec)

    # Loose recovery: some files have a tail chunk with multiple records concatenated without tabs/newlines.
    # We'll recover "URL + nearby company name" patterns for URLs not already captured as entrypoints.
    captured_urls = set(it.get("recruitment_url") for it in out if it.get("recruitment_url"))
    all_urls = list(dict.fromkeys(re.findall(_ASCII_URL_RE, t)))

    majors = sorted({(it.get("_industry_major") or "") for it in out if it.get("_industry_major")}, key=len, reverse=True)
    tracks = []
    for it in out:
        ind = (it.get("industry") or "")
        if "/" in ind:
            tracks.append(ind.split("/", 1)[1].strip())
    tracks = sorted({x for x in tracks if x and len(x) <= 20}, key=len, reverse=True)

    # Company/org suffixes used for name inference in concatenated chunks.
    # Keep the primary list reasonably "strong": avoid single-character tokens (e.g. "院/所"),
    # and keep very generic tokens (e.g. "...中心") as a weak fallback to reduce false positives.
    strong_suffixes = [
        "股份有限公司",
        "有限责任公司",
        "股份公司",
        "有限公司",
        "集团",
        "控股集团",
        "控股",
        "研究院",
        "研究所",
        "实验室",
        "委员会",
        "总行",
        "银行",
        "交易所",
        "出版社",
        "清算有限公司",
        "总公司",
        "股份",
        "公司",
    ]
    weak_suffixes = [
        "中心",
    ]

    # Note: avoid '/' in these patterns, or you may accidentally capture "https://.../央企国企..." blobs.
    strong_suffixes_sorted = sorted({s for s in strong_suffixes if s}, key=len, reverse=True)
    weak_suffixes_sorted = sorted({s for s in weak_suffixes if s}, key=len, reverse=True)

    suffix_re_strong = re.compile(
        r"([A-Za-z0-9\u4e00-\u9fff·（）() &.-]{2,80}?(?:"
        + "|".join(map(re.escape, strong_suffixes_sorted))
        + r"))"
    )
    suffix_re_weak = (
        re.compile(
            r"([A-Za-z0-9\u4e00-\u9fff·（）() &.-]{2,80}?(?:"
            + "|".join(map(re.escape, weak_suffixes_sorted))
            + r"))"
        )
        if weak_suffixes_sorted
        else None
    )
    paren_re = re.compile(r"([A-Za-z0-9\u4e00-\u9fff·（）() &.-]{2,80}?[\(（][^()\n]{1,60}[\)）])")

    def _guess_inline_name(context: str) -> str | None:
        ctx = (context or "").replace("\t", " ").replace("\n", " ")
        # Remove any URLs that belong to previous records in the concatenated chunk.
        ctx = re.sub(_ASCII_URL_RE, " ", ctx)
        ctx = re.sub(r"\s+", " ", ctx).strip()
        if not ctx:
            return None

        common_track_prefixes = [
            "国家金融基建",
            "证券与交易",
            "信创与云计算",
            "交通与航空",
            "工业互联网",
            "医疗科技与生信",
            "智能家电与IoT",
            "高校系硬科技",
            "学术与出版",
            "超算与大数据",
            "金融云与AI",
        ]

        # Prefer "...(Alias)" patterns near the end.
        # 160 chars is often too short for concatenated records with long focus text.
        tail = ctx[-360:] if len(ctx) > 360 else ctx
        short_pat = (
            r"([A-Za-z0-9\u4e00-\u9fff·&.-]{2,30}(?:"
            + "|".join(map(re.escape, strong_suffixes_sorted))
            + r"))"
        )

        m_all = list(paren_re.finditer(tail))
        if m_all:
            # Take the rightmost match (closest to the URL we are parsing).
            cand = m_all[-1].group(1).strip()
        else:
            # No alias pattern: try an overlapping suffix token first (works well when there are no separators).
            short_hits0 = list(re.finditer(r"(?=" + short_pat + r")", tail))
            if short_hits0:
                cands0 = [m.group(1).strip() for m in short_hits0 if m.group(1)]
                prefer_longer_suffixes = ("出版社", "研究院", "研究所", "实验室", "委员会", "交易所", "银行")
                if any(c.endswith(prefer_longer_suffixes) for c in cands0):
                    cands2 = [c for c in cands0 if c.endswith(prefer_longer_suffixes)]
                    cand = max(cands2, key=len) if cands2 else cands0[-1]
                else:
                    # Rightmost match tends to drop category prefixes like "...微电网", keeping the org name token.
                    cand = cands0[-1]
            else:
                # Fall back to regex matches (non-overlapping).
                s_all = list(suffix_re_strong.finditer(tail))
                if s_all:
                    cand = s_all[-1].group(1).strip()
                elif suffix_re_weak:
                    w_all = list(suffix_re_weak.finditer(tail))
                    cand = w_all[-1].group(1).strip() if w_all else ""
                else:
                    cand = ""
        if not cand:
            return None

        # Reduce to a plausible company name.
        # 1) If we have "(Alias)" take the head part, then keep the last short "name-like" phrase.
        head = cand
        for sep in ["（", "("]:
            if sep in head:
                head = head.split(sep, 1)[0].strip()
                break

        # Drop leading major/track tokens if they got glued in front.
        # Some concatenated chunks have a leading "等/、" before the major, so we accept near-start matches.
        head = head.lstrip(" 等、,，;；:：|/\\-")
        changed = True
        while changed:
            changed = False
            for maj in majors:
                if not maj:
                    continue
                pos = head.find(maj)
                if pos != -1 and pos <= 3:
                    head = head[pos + len(maj) :].strip()
                    head = head.lstrip(" 等、,，;；:：|/\\-")
                    changed = True
            for tr in tracks:
                if not tr:
                    continue
                pos = head.find(tr)
                if pos != -1 and pos <= 3:
                    head = head[pos + len(tr) :].strip()
                    head = head.lstrip(" 等、,，;；:：|/\\-")
                    changed = True
            for tp in common_track_prefixes:
                pos = head.find(tp)
                if pos != -1 and pos <= 3:
                    head = head[pos + len(tp) :].strip()
                    head = head.lstrip(" 等、,，;；:：|/\\-")
                    changed = True
        head = re.sub(r"^[\u4e00-\u9fff]{2,10}与[\u4e00-\u9fffA-Za-z]{1,10}", "", head).strip()

        # Prefer a short suffix-based company phrase if present.
        # Use an overlapping matcher so we can pick the rightmost name-like token even when the string has no separators.
        short_hits = list(re.finditer(r"(?=" + short_pat + r")", head))
        if short_hits:
            cands = [m.group(1).strip() for m in short_hits if m.group(1)]
            # For some suffix types, the rightmost match can be too short (e.g. "教育出版社" vs "高等教育出版社").
            prefer_longer_suffixes = ("出版社", "研究院", "研究所", "实验室", "委员会", "交易所", "银行")
            if any(c.endswith(prefer_longer_suffixes) for c in cands):
                cands2 = [c for c in cands if c.endswith(prefer_longer_suffixes)]
                cand2 = max(cands2, key=len) if cands2 else cands[-1]
            else:
                # Rightmost match tends to drop category prefixes like "...微电网", keeping the org name token.
                cand2 = cands[-1]
        else:
            # Fallback: last contiguous token.
            m = re.search(r"([A-Za-z0-9\u4e00-\u9fff·&.-]{2,40})$", head)
            cand2 = m.group(1).strip() if m else head.strip()

        cand = cand2.strip(" -_/|·")

        cand = re.sub(r"\s+", " ", cand).strip()
        if len(cand) < 2 or len(cand) > 80:
            return None
        return cand

    def _guess_inline_major(context: str) -> str:
        # Find last major keyword occurrence in the context.
        ctx = context or ""
        best = ""
        best_pos = -1
        for maj in majors:
            if not maj:
                continue
            pos = ctx.rfind(maj)
            if pos > best_pos:
                best_pos = pos
                best = maj
        return best or ""

    # Use URL position boundaries to avoid mixing contexts from previous URLs in concatenated chunks.
    url_matches = list(re.finditer(_ASCII_URL_RE, t))
    prev_end = 0
    for m in url_matches:
        url = m.group(0)
        idx = m.start()

        # Always advance prev_end so the segmentation is stable.
        seg = t[max(0, prev_end) : idx]
        prev_end = m.end()

        if url in captured_urls:
            continue

        # Prefer the chunk since the previous URL. If it contains newlines, only keep the last line,
        # because that's where the fields for the current URL typically are.
        left = seg
        if "\n" in left:
            left = left[left.rfind("\n") + 1 :]
        # Use a bigger window for glued tail chunks (260 chars is often too short to capture the org name).
        left = left[-900:] if len(left) > 900 else left

        name = _guess_inline_name(left)
        if not name:
            # Fallback: broader lookback if the segment is tiny (e.g., the first URL in the glued tail).
            left2 = t[max(0, idx - 1400) : idx]
            name = _guess_inline_name(left2)
            if not name:
                continue

        major = _guess_inline_major(left)
        rec = _row(major or "", None, name, None, url)
        if rec:
            out.append(rec)

    # Dedup by name, keep the one with a longer URL (often more specific).
    best: dict[str, dict] = {}
    for it in out:
        n = it["name"]
        prev = best.get(n)
        if not prev:
            best[n] = it
            continue
        if len(str(it.get("recruitment_url") or "")) > len(str(prev.get("recruitment_url") or "")):
            best[n] = it

    return list(best.values())


def _m_zhiye_base(url: str) -> str | None:
    u = (url or "").strip()
    if not u:
        return None
    try:
        from urllib.parse import urlparse

        p = urlparse(u)
        host = (p.netloc or "").strip()
        if not host:
            return None
        if host.endswith(".m.zhiye.com"):
            return f"{p.scheme or 'https'}://{host}"
        if host.endswith(".zhiye.com"):
            sub = host[: -len(".zhiye.com")]
            if sub:
                return f"{p.scheme or 'https'}://{sub}.m.zhiye.com"
    except Exception:
        return None
    return None


def _is_hotjob(url: str) -> bool:
    u = (url or "").strip()
    if not u:
        return False
    try:
        from urllib.parse import urlparse

        p = urlparse(u)
        host = (p.netloc or "").lower().strip()
    except Exception:
        return False
    return host.endswith(".hotjob.cn") or host == "hotjob.cn"


def cmd_import_companies_tsv(args: argparse.Namespace) -> None:
    """Import companies from a local TSV text file and create official crawl sources.

    This is designed for internal usage: you keep a flat text list of companies
    and their official entrypoints, then this command upserts companies and
    creates/updates `Official:<company>` CrawlSource entries.
    """

    proxy = (str(args.proxy or "").strip() or None)
    text = _read_text_best_effort(str(args.file))
    items = _parse_company_tsv_text(text)
    if not items:
        raise SystemExit("no companies parsed from file (expected TSV with URLs)")

    disable_global = bool(args.disable_global)
    force_entrypoint = bool(args.force_entrypoint)

    title_contains = [
        "后端",
        "前端",
        "全栈",
        "开发",
        "工程师",
        "架构",
        "数据",
        "算法",
        "测试",
        "运维",
        "DevOps",
        "SRE",
        "金融科技",
        "银行",
        "新能源",
        "储能",
        "锂电",
        "电池",
        "电化学",
        "材料",
        "化工",
        "研发",
        "项目",
        "项目管理",
    ]
    url_contains = ["job", "jobs", "career", "careers", "recruit", "zhaopin", "hr", "join"]
    url_excludes = ["campus", "intern", "xiaozhao", "校园", "校招", "实习"]

    db = SessionLocal()
    try:
        c_created = 0
        c_updated = 0
        src_created = 0
        src_updated = 0
        disabled = 0

        for it in sorted(items, key=lambda x: x.get("name") or ""):
            name = str(it.get("name") or "").strip()
            rec_url = str(it.get("recruitment_url") or "").strip()
            if not name:
                continue

            c = db.execute(select(Company).where(Company.name == name)).scalar_one_or_none()
            if not c:
                c = Company(name=name)
                db.add(c)
                db.flush()
                c_created += 1
            else:
                c_updated += 1

            if not c.industry and it.get("industry"):
                c.industry = str(it["industry"]).strip()
            if not c.company_type and it.get("company_type"):
                c.company_type = str(it["company_type"]).strip()
            if not c.focus_directions and it.get("focus_directions"):
                c.focus_directions = str(it["focus_directions"]).strip()

            if rec_url:
                if force_entrypoint or not (c.recruitment_url and str(c.recruitment_url).strip()):
                    c.recruitment_url = rec_url

            db.add(c)

            if not rec_url:
                continue

            src_name = f"Official:{name}"
            enabled = True
            industry_major = str(it.get("_industry_major") or "")
            if disable_global and ("全球" in industry_major or "跨国" in industry_major):
                enabled = False
                disabled += 1

            # Avoid auto-crawling large job platforms that typically require interactive search/login and have strict ToS.
            # Users can still use /jobs/import (URL prefill) or url_list for explicit job detail URLs.
            try:
                from urllib.parse import urlparse

                host = (urlparse(rec_url).netloc or "").lower().strip()
            except Exception:
                host = ""
            platform_block = [
                "www.zhipin.com",  # Boss直聘
                "zhipin.com",
                "jobs.51job.com",
                "www.51job.com",
                "51job.com",
                "www.indeed.com",
                "indeed.com",
                "www.zhaopin.com",  # 智联
                "zhaopin.com",
                "www.liepin.com",
                "liepin.com",
                "zhaopin.58.com",
                "58.com",
            ]
            if host and any(host == h or host.endswith("." + h) for h in platform_block):
                if enabled:
                    enabled = False
                    disabled += 1

            kind = "html_list"
            cfg: dict = {}

            mz_base = _m_zhiye_base(rec_url)
            if mz_base:
                kind = "m_zhiye"
                cfg = {"base_url": mz_base, "company_name": name, "jc": 1, "page_size": 30, "max_pages": 40, "source_type": "official"}
            elif _is_hotjob(rec_url):
                try:
                    from urllib.parse import urlparse

                    p = urlparse(rec_url)
                    base = f"{p.scheme or 'https'}://{p.netloc}" if p.netloc else rec_url
                except Exception:
                    base = rec_url
                kind = "hotjob"
                cfg = {"base_url": base.replace("http://", "https://"), "company_name": name, "recruit_type": 2, "page_size": 12, "max_pages": 12, "source_type": "official"}
            else:
                cfg = {
                    "list_url": rec_url,
                    "company_name": name,
                    "url_contains": url_contains,
                    "url_excludes": url_excludes,
                    "title_contains": title_contains,
                    "max_items": int(args.max_items),
                    "source_type": "official",
                }

            if proxy:
                cfg["proxy"] = proxy

            existing = db.execute(select(CrawlSource).where(CrawlSource.name == src_name)).scalar_one_or_none()
            if existing:
                existing.kind = kind
                existing.enabled = bool(enabled)
                existing.config = cfg
                db.add(existing)
                src_updated += 1
            else:
                db.add(CrawlSource(kind=kind, name=src_name, enabled=bool(enabled), config=cfg))
                src_created += 1

        db.commit()
        print(
            "imported companies from tsv: "
            f"companies_created={c_created} companies_updated={c_updated} "
            f"sources_created={src_created} sources_updated={src_updated} disabled={disabled} total_parsed={len(items)}"
        )
    finally:
        db.close()


def cmd_import_companies(args: argparse.Namespace) -> None:
    """Upsert companies (name-unique) from a JSON file."""

    data = _load_json(args.file)
    if not isinstance(data, list):
        raise SystemExit("companies file must be a JSON array")

    db = SessionLocal()
    try:
        created = 0
        updated = 0

        for item in data:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name") or "").strip()
            if not name:
                continue

            c = db.execute(select(Company).where(Company.name == name)).scalar_one_or_none()
            if not c:
                c = Company(name=name)
                db.add(c)
                db.flush()
                created += 1
            else:
                updated += 1

            # Only set fields when present in input.
            if "industry" in item and (item.get("industry") is None or str(item.get("industry")).strip()):
                c.industry = (str(item.get("industry")).strip() if item.get("industry") is not None else None)
            if "hq_location" in item and (item.get("hq_location") is None or str(item.get("hq_location")).strip()):
                c.hq_location = (str(item.get("hq_location")).strip() if item.get("hq_location") is not None else None)
            if "focus_directions" in item and (item.get("focus_directions") is None or str(item.get("focus_directions")).strip()):
                c.focus_directions = (
                    (str(item.get("focus_directions")).strip() if item.get("focus_directions") is not None else None)
                )
            if "company_type" in item and (item.get("company_type") is None or str(item.get("company_type")).strip()):
                c.company_type = (str(item.get("company_type")).strip() if item.get("company_type") is not None else None)
            if "website" in item and (item.get("website") is None or str(item.get("website")).strip()):
                c.website = (str(item.get("website")).strip() if item.get("website") is not None else None)
            if "recruitment_url" in item and (item.get("recruitment_url") is None or str(item.get("recruitment_url")).strip()):
                c.recruitment_url = (str(item.get("recruitment_url")).strip() if item.get("recruitment_url") is not None else None)

            db.add(c)

        db.commit()
        print(f"imported companies: created={created} updated={updated} total={created+updated}")
    finally:
        db.close()


def _infer_industry(name: str, note: str) -> str | None:
    text = f"{name} {note}".lower()
    if any(k in text for k in ["电池", "锂电", "储能", "bms", "电化学", "新能源"]):
        return "电池与新能源"
    if any(k in text for k in ["电网", "电力", "发电", "华能", "华电", "国家电网", "南方电网", "三峡"]):
        return "能源与电力"
    if any(k in text for k in ["银行", "金科", "金融科技", "fintech", "核心系统"]):
        return "银行与金融科技"
    if any(k in text for k in ["研究院", "研究所", "科学院"]):
        return "科技与软件"
    return None


def cmd_import_companies_xlsx(args: argparse.Namespace) -> None:
    """Import companies from an Excel file that contains at least a '单位' column.

    This supports the user's internal delivery tracking sheet:
    columns: 分类/投递日期/单位/地点/级别/备注

    Behavior:
    - Upsert Company by name.
    - Optionally create per-company CrawlSource entries (Guopin keyword search) as a coverage booster.
    """

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise SystemExit(f"missing dependency openpyxl: {e}")

    xlsx_path = str(args.file)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    sheets: list[str]
    if args.sheet:
        sheets = [args.sheet]
    else:
        # Default: only enterprise-oriented sheets (skip civil-service "选调" sheet).
        sheets = [sn for sn in wb.sheetnames if "选调" not in sn]

    db = SessionLocal()
    try:
        created = 0
        updated = 0
        units: dict[str, dict] = {}

        for sn in sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]

            # Expect header row in row 1.
            header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            col = {name: idx for idx, name in enumerate(header) if name}
            if "单位" not in col:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                name = row[col["单位"]] if col["单位"] < len(row) else None
                if name is None:
                    continue
                name = str(name).strip()
                if not name:
                    continue
                # Skip obvious non-company numeric placeholders.
                if name.isdigit():
                    continue

                city = None
                if "地点" in col and col["地点"] < len(row):
                    v = row[col["地点"]]
                    if v is not None and str(v).strip():
                        city = str(v).strip()

                level = None
                if "级别" in col and col["级别"] < len(row):
                    v = row[col["级别"]]
                    if v is not None and str(v).strip():
                        level = str(v).strip()

                note = None
                if "备注" in col and col["备注"] < len(row):
                    v = row[col["备注"]]
                    if v is not None and str(v).strip():
                        note = str(v).strip()

                industry = _infer_industry(name, note or "")

                # Keep the "best" info if a company appears multiple times.
                agg = units.get(name) or {}
                if level and not agg.get("company_type"):
                    agg["company_type"] = level
                if city and not agg.get("hq_location"):
                    agg["hq_location"] = city
                if industry and not agg.get("industry"):
                    agg["industry"] = industry
                if note and not agg.get("focus_directions"):
                    # Keep notes short; treat as focus directions hint.
                    agg["focus_directions"] = note[:200]
                units[name] = agg

        for name, meta in sorted(units.items(), key=lambda kv: kv[0]):
            c = db.execute(select(Company).where(Company.name == name)).scalar_one_or_none()
            if not c:
                c = Company(name=name)
                db.add(c)
                db.flush()
                created += 1
            else:
                updated += 1

            # Only fill empty fields; do not overwrite curated seed data.
            if not c.company_type and meta.get("company_type"):
                c.company_type = meta["company_type"]
            if not c.hq_location and meta.get("hq_location"):
                c.hq_location = meta["hq_location"]
            if not c.industry and meta.get("industry"):
                c.industry = meta["industry"]
            if not c.focus_directions and meta.get("focus_directions"):
                c.focus_directions = meta["focus_directions"]

            db.add(c)

        db.commit()

        src_created = 0
        src_updated = 0
        if args.create_sources:
            include = [
                "新能源",
                "锂电",
                "电池",
                "电芯",
                "储能",
                "BMS",
                "电化学",
                "材料",
                "研发",
                "后端",
                "后台",
                "软件开发",
                "系统开发",
                "全栈",
                "大数据",
                "数据平台",
                "数据工程",
                "数据开发",
                "数据分析",
                "算法",
                "人工智能",
                "平台",
                "中台",
                "云",
                "微服务",
                "DevOps",
                "SRE",
                "测试",
                "银行科技",
                "金融科技",
                "项目管理",
                "架构",
                "架构师",
            ]
            exclude = ["销售", "市场", "商务", "运营", "客服", "财务", "法务", "人力", "行政"]

            city_allow = ["北京", "上海", "广州", "深圳"] if args.city_allowlist else []

            for name in sorted(units.keys()):
                src_name = f"Guopin:{name}"
                cfg = {
                    "company_name": "国聘网",
                    "api_base": "https://gp-api.iguopin.com",
                    "keyword": name,
                    "page_size": int(args.page_size),
                    "max_pages": int(args.max_pages),
                    "include_keywords": include,
                    "exclude_keywords": exclude,
                    "source_type": "aggregator",
                }
                if city_allow:
                    cfg["city_allowlist"] = city_allow

                existing = db.execute(select(CrawlSource).where(CrawlSource.name == src_name)).scalar_one_or_none()
                if existing:
                    existing.kind = "iguopin"
                    existing.enabled = True
                    existing.config = cfg
                    db.add(existing)
                    src_updated += 1
                else:
                    s = CrawlSource(kind="iguopin", name=src_name, enabled=True, config=cfg)
                    db.add(s)
                    src_created += 1

            db.commit()

        print(
            f"imported companies from xlsx: companies_created={created} companies_updated={updated} "
            f"sources_created={src_created} sources_updated={src_updated} total_companies={len(units)}"
        )
    finally:
        db.close()
        try:
            wb.close()
        except Exception:
            pass


def cmd_seed_official_html_sources(args: argparse.Namespace) -> None:
    """Create html_list crawl sources for companies that have recruitment_url set.

    This is best-effort: many official portals are JS-heavy or WAF-protected.
    """

    proxy = str(args.proxy or "").strip() or None
    title_contains = [
        "后端",
        "前端",
        "全栈",
        "开发",
        "工程师",
        "架构",
        "数据",
        "算法",
        "测试",
        "运维",
        "DevOps",
        "SRE",
        "金融科技",
        "银行",
        "新能源",
        "储能",
        "锂电",
        "电池",
        "电化学",
        "材料",
        "化工",
        "研发",
        "项目",
        "项目管理",
    ]
    url_contains = ["job", "jobs", "career", "careers", "recruit", "zhaopin", "hr", "join"]
    url_excludes = ["campus", "intern", "xiaozhao", "校园", "校招", "实习"]

    db = SessionLocal()
    try:
        companies = (
            db.execute(
                select(Company)
                .where(Company.recruitment_url.is_not(None))
                .where(Company.recruitment_url != "")
                .order_by(Company.name.asc())
            )
            .scalars()
            .all()
        )

        def _m_zhiye_base(url: str) -> str | None:
            u = (url or "").strip()
            if not u:
                return None
            # Common pattern: https://xxx.zhiye.com/... -> https://xxx.m.zhiye.com
            try:
                from urllib.parse import urlparse

                p = urlparse(u)
                host = (p.netloc or "").strip()
                if not host:
                    return None
                if host.endswith(".m.zhiye.com"):
                    return f"{p.scheme}://{host}"
                if host.endswith(".zhiye.com"):
                    sub = host[: -len(".zhiye.com")]
                    if sub:
                        return f"{p.scheme}://{sub}.m.zhiye.com"
            except Exception:
                return None
            return None

        def _is_hotjob(url: str) -> bool:
            u = (url or "").strip().lower()
            if not u:
                return False
            try:
                from urllib.parse import urlparse

                p = urlparse(u)
                host = (p.netloc or "").lower()
            except Exception:
                return False
            return host.endswith(".hotjob.cn") or host == "hotjob.cn"

        created = 0
        updated = 0
        for c in companies:
            name = c.name
            src_name = f"Official:{name}"
            kind = "html_list"
            cfg: dict = {}

            rec_url = (c.recruitment_url or "").strip()
            mz_base = _m_zhiye_base(rec_url) if rec_url else None
            if mz_base:
                kind = "m_zhiye"
                cfg = {
                    "base_url": mz_base,
                    "company_name": name,
                    "jc": 1,  # 社招
                    "page_size": 30,
                    "max_pages": 30,
                    "source_type": "official",
                }
            elif _is_hotjob(rec_url):
                # Hotjob (大易/微招聘) official portal, use wecruit public endpoints.
                kind = "hotjob"
                # Normalize to origin only. (http -> https is fine)
                try:
                    from urllib.parse import urlparse

                    p = urlparse(rec_url)
                    base = f"{p.scheme or 'https'}://{p.netloc}" if p.netloc else rec_url
                except Exception:
                    base = rec_url
                cfg = {
                    "base_url": base.replace("http://", "https://"),
                    "company_name": name,
                    "recruit_type": 2,  # 社招
                    "page_size": 12,
                    "max_pages": 12,
                    "source_type": "official",
                }
            else:
                cfg = {
                    "list_url": c.recruitment_url,
                    "company_name": name,
                    "url_contains": url_contains,
                    "url_excludes": url_excludes,
                    "title_contains": title_contains,
                    "max_items": int(args.max_items),
                    "source_type": "official",
                }

            if proxy:
                cfg["proxy"] = proxy

            existing = db.execute(select(CrawlSource).where(CrawlSource.name == src_name)).scalar_one_or_none()
            if existing:
                existing.kind = kind
                existing.enabled = True
                existing.config = cfg
                db.add(existing)
                updated += 1
            else:
                s = CrawlSource(kind=kind, name=src_name, enabled=True, config=cfg)
                db.add(s)
                created += 1

        db.commit()
        print(f"seeded official html sources: created={created} updated={updated} total={created+updated}")
    finally:
        db.close()


def cmd_backfill_job_source_kind(args: argparse.Namespace) -> None:
    """Best-effort backfill for JobSource.source_kind on existing data."""

    db = SessionLocal()
    try:
        rows = db.execute(select(JobSource).where(JobSource.source_kind.is_(None))).scalars().all()
        updated = 0
        for js in rows:
            name = (js.source_name or "").strip()
            st = (js.source_type or "").strip()

            kind = None
            if st == "import":
                kind = "import"
            elif name.lower() == "tencent":
                kind = "tencent"
            elif name.lower() == "kuaishou":
                kind = "kuaishou"
            elif name.lower() == "jd":
                kind = "jd"
            elif name.lower().startswith("guopin"):
                kind = "iguopin"
            elif name.lower().startswith("official:"):
                kind = "html_list"

            if kind:
                js.source_kind = kind
                db.add(js)
                updated += 1

        db.commit()
        print(f"backfilled job_sources.source_kind: updated={updated} scanned={len(rows)}")
    finally:
        db.close()


def main() -> None:
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="cmd", required=True)

    cu = sub.add_parser("create-user")
    cu.add_argument("--username", required=True)
    cu.add_argument("--password", required=True)
    cu.add_argument("--admin", action="store_true")
    cu.set_defaults(fn=cmd_create_user)

    sp = sub.add_parser("set-password")
    sp.add_argument("--username", required=True)
    sp.add_argument("--password", required=True)
    sp.set_defaults(fn=cmd_set_password)

    eu = sub.add_parser("ensure-user")
    eu.add_argument("--username", required=True)
    eu.add_argument("--password", required=True)
    eu.add_argument("--admin", action="store_true")
    eu.set_defaults(fn=cmd_ensure_user)

    ic = sub.add_parser("import-companies")
    ic.add_argument("--file", required=True, help="JSON array of {name, industry, company_type, recruitment_url, website}")
    ic.set_defaults(fn=cmd_import_companies)

    ix = sub.add_parser("import-companies-xlsx")
    ix.add_argument("--file", required=True, help="Excel file containing a '单位' column")
    ix.add_argument("--sheet", default="", help="Sheet name (default: all sheets)")
    ix.add_argument("--create-sources", action="store_true", help="Create per-company Guopin crawl sources (keyword search)")
    ix.add_argument("--max-pages", type=int, default=2)
    ix.add_argument("--page-size", type=int, default=50)
    ix.add_argument("--city-allowlist", action="store_true", help="Restrict ingestion to 北上广深 (based on city field)")
    ix.set_defaults(fn=cmd_import_companies_xlsx)

    it = sub.add_parser("import-companies-tsv")
    it.add_argument("--file", required=True, help="TSV text file with columns: 行业大类/细分赛道/企业/重点方向/官方入口(URL)")
    it.add_argument("--proxy", default="", help="Optional proxy URL, e.g. http://127.0.0.1:7890")
    it.add_argument("--max-items", type=int, default=200)
    it.add_argument("--disable-global", action="store_true", help="Disable crawl sources for '全球跨国巨头' rows (keeps domestic focus)")
    it.add_argument("--force-entrypoint", action="store_true", help="Overwrite existing recruitment_url if present")
    it.set_defaults(fn=cmd_import_companies_tsv)

    so = sub.add_parser("seed-official-html-sources")
    so.add_argument("--proxy", default="", help="Optional proxy URL, e.g. http://127.0.0.1:7890")
    so.add_argument("--max-items", type=int, default=200)
    so.set_defaults(fn=cmd_seed_official_html_sources)

    bf = sub.add_parser("backfill-job-source-kind")
    bf.set_defaults(fn=cmd_backfill_job_source_kind)

    args = p.parse_args()
    args.fn(args)


if __name__ == "__main__":
    main()
