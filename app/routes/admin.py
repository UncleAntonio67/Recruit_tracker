from __future__ import annotations

import os
import json
import re
from io import BytesIO

from datetime import timedelta

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.auth import require_admin
from app.db import get_db
from app.models import Company, CrawlSource, User
from app.security import hash_password
from app.views import templates

router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/users", response_class=HTMLResponse)
def users_list(
    request: Request,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> HTMLResponse:
    users = db.execute(select(User).order_by(User.created_at.desc())).scalars().all()
    return templates.TemplateResponse("admin_users.html", {"request": request, "user": admin, "users": users})


@router.get("/users/new", response_class=HTMLResponse)
def user_new_page(
    request: Request,
    admin: User = Depends(require_admin),
) -> HTMLResponse:
    return templates.TemplateResponse("admin_user_new.html", {"request": request, "user": admin})


@router.post("/users/new")
def user_new_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    is_admin: str | None = Form(default=None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> RedirectResponse:
    u = db.execute(select(User).where(User.username == username.strip())).scalar_one_or_none()
    if u:
        return templates.TemplateResponse(
            "admin_user_new.html",
            {"request": request, "user": admin, "error": "用户名已存在"},
            status_code=400,
        )

    new_u = User(
        username=username.strip(),
        password_hash=hash_password(password),
        is_admin=bool(is_admin),
    )
    db.add(new_u)
    db.commit()
    return RedirectResponse(url="/admin/users", status_code=302)


@router.get("/sources", response_class=HTMLResponse)
def sources_list(
    request: Request,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> HTMLResponse:
    sources = db.execute(select(CrawlSource).order_by(CrawlSource.created_at.desc())).scalars().all()
    crawl_interval_hours = int((os.environ.get("CRAWL_INTERVAL_HOURS") or "0").strip() or 0)
    items = []
    for s in sources:
        next_run_at = None
        if crawl_interval_hours > 0 and s.last_run_at is not None:
            try:
                next_run_at = s.last_run_at + timedelta(hours=crawl_interval_hours)
            except Exception:
                next_run_at = None
        items.append({"source": s, "next_run_at": next_run_at})
    return templates.TemplateResponse(
        "admin_sources.html",
        {
            "request": request,
            "user": admin,
            "items": items,
            "crawl_interval_hours": crawl_interval_hours,
            "crawl_mode": (os.environ.get("CRAWL_MODE") or "official").strip() or "official",
            "crawl_since_days": int((os.environ.get("CRAWL_SINCE_DAYS") or "180").strip() or 180),
        },
    )


@router.get("/sources/new", response_class=HTMLResponse)
def sources_new_page(
    request: Request,
    admin: User = Depends(require_admin),
) -> HTMLResponse:
    kinds = ["tencent", "kuaishou", "iguopin", "jd", "m_zhiye", "hotjob", "greenhouse", "lever", "rss", "html_list", "url_list"]
    default_config = '{\n  "company_name": "",\n  "...": ""\n}'
    return templates.TemplateResponse(
        "admin_source_new.html",
        {"request": request, "user": admin, "kinds": kinds, "default_config": default_config},
    )


def _infer_official_source(company_name: str, entry_url: str, *, proxy: str | None) -> tuple[str, dict]:
    """Infer a CrawlSource kind+config from an official entrypoint URL."""

    from urllib.parse import urlparse

    u = entry_url.strip()
    p = urlparse(u)
    host = (p.netloc or "").strip().lower()

    kind = "html_list"
    cfg: dict = {
        "list_url": u,
        "company_name": company_name.strip(),
        "url_contains": ["job", "jobs", "career", "careers", "recruit", "zhaopin", "hr", "join"],
        "url_excludes": ["campus", "intern", "xiaozhao", "校园", "校招", "实习"],
        "max_items": 200,
        "source_type": "official",
    }

    # Beisen Zhiye portal: use structured API connector.
    if host.endswith(".m.zhiye.com"):
        kind = "m_zhiye"
        cfg = {"base_url": f"{p.scheme}://{p.netloc}", "company_name": company_name.strip(), "jc": 1, "page_size": 30, "max_pages": 40, "source_type": "official"}
    elif host.endswith(".zhiye.com"):
        sub = host[: -len(".zhiye.com")]
        if sub:
            kind = "m_zhiye"
            cfg = {
                "base_url": f"{p.scheme}://{sub}.m.zhiye.com",
                "company_name": company_name.strip(),
                "jc": 1,
                "page_size": 30,
                "max_pages": 40,
                "source_type": "official",
            }
    elif host.endswith(".hotjob.cn") or host == "hotjob.cn":
        kind = "hotjob"
        base = f"{p.scheme or 'https'}://{p.netloc}" if p.netloc else u
        cfg = {
            "base_url": base.replace("http://", "https://"),
            "company_name": company_name.strip(),
            "recruit_type": 2,  # 社招
            "page_size": 12,
            "max_pages": 12,
            "source_type": "official",
        }
    elif u.lower().endswith((".rss", ".xml")):
        kind = "rss"
        cfg = {"feed_url": u, "company_name": company_name.strip(), "source_type": "official"}

    if proxy and proxy.strip():
        cfg["proxy"] = proxy.strip()
    return kind, cfg


@router.post("/sources/new_simple")
def sources_new_simple_post(
    request: Request,
    company_name: str = Form(...),
    entry_url: str = Form(...),
    name: str | None = Form(default=None),
    proxy: str | None = Form(default=None),
    enabled: str | None = Form(default=None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> RedirectResponse:
    cn = company_name.strip()
    eu = entry_url.strip()
    if not cn or not eu:
        return RedirectResponse(url="/admin/sources/new", status_code=302)

    src_name = (name or "").strip() or f"Official:{cn}"
    exists = db.execute(select(CrawlSource).where(CrawlSource.name == src_name)).scalar_one_or_none()
    if exists:
        kinds = ["tencent", "kuaishou", "iguopin", "jd", "m_zhiye", "hotjob", "greenhouse", "lever", "rss", "html_list", "url_list"]
        return templates.TemplateResponse(
            "admin_source_new.html",
            {
                "request": request,
                "user": admin,
                "kinds": kinds,
                "error": "名称已存在(简单模式)。请改名或在公司页点击“生成/更新采集源”。",
                "simple": {"company_name": cn, "entry_url": eu, "name": src_name, "proxy": (proxy or "")},
            },
            status_code=400,
        )

    kind, cfg = _infer_official_source(cn, eu, proxy=proxy)
    s = CrawlSource(kind=kind, name=src_name, enabled=bool(enabled), config=cfg)
    db.add(s)
    db.commit()
    return RedirectResponse(url="/admin/sources", status_code=302)


@router.post("/sources/new")
def sources_new_post(
    request: Request,
    kind: str = Form(...),
    name: str = Form(...),
    enabled: str | None = Form(default=None),
    config_json: str = Form(...),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> RedirectResponse:
    try:
        cfg = json.loads(config_json) if config_json.strip() else {}
    except Exception:
        kinds = ["tencent", "kuaishou", "iguopin", "jd", "m_zhiye", "hotjob", "greenhouse", "lever", "rss", "html_list", "url_list"]
        return templates.TemplateResponse(
            "admin_source_new.html",
            {"request": request, "user": admin, "kinds": kinds, "error": "config_json 不是合法 JSON", "config_json": config_json},
            status_code=400,
        )

    exists = db.execute(select(CrawlSource).where(CrawlSource.name == name.strip())).scalar_one_or_none()
    if exists:
        kinds = ["tencent", "kuaishou", "iguopin", "jd", "m_zhiye", "hotjob", "greenhouse", "lever", "rss", "html_list", "url_list"]
        return templates.TemplateResponse(
            "admin_source_new.html",
            {"request": request, "user": admin, "kinds": kinds, "error": "名称已存在", "config_json": config_json},
            status_code=400,
        )

    s = CrawlSource(kind=kind.strip(), name=name.strip(), enabled=bool(enabled), config=cfg)
    db.add(s)
    db.commit()
    return RedirectResponse(url="/admin/sources", status_code=302)


@router.get("/companies/import", response_class=HTMLResponse)
def companies_import_page(
    request: Request,
    admin: User = Depends(require_admin),
) -> HTMLResponse:
    return templates.TemplateResponse("admin_companies_import.html", {"request": request, "user": admin})


def _upsert_company(db: Session, name: str) -> tuple[Company, bool]:
    n = name.strip()
    existing = db.execute(select(Company).where(Company.name == n)).scalar_one_or_none()
    if existing:
        return existing, False
    c = Company(name=n)
    db.add(c)
    db.flush()
    return c, True


def _parse_companies_from_txt(blob: bytes) -> list[dict]:
    # Try utf-8 first, then gbk for Windows-exported files.
    text = None
    for enc in ("utf-8", "utf-8-sig", "gbk"):
        try:
            text = blob.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        return []

    url_re = re.compile(r"https?://\\S+", flags=re.IGNORECASE)
    rows: list[dict] = []

    for raw in text.splitlines():
        line = (raw or "").strip()
        if not line:
            continue
        if line.startswith("#"):
            continue

        # Prefer tab-separated: name \t ... \t url
        parts = [p.strip() for p in line.split("\\t") if p.strip()]
        url = None
        m = url_re.search(line)
        if m:
            url = m.group(0).rstrip("),.;]")  # minor cleanup

        name = ""
        if len(parts) >= 2 and url and parts[-1].startswith("http"):
            name = parts[0]
            url = parts[-1]
        elif url:
            name = line.split(url, 1)[0].strip(" -|:：\u3000")

        if not name:
            continue

        rows.append({"name": name[:120], "recruitment_url": url})

    return rows


def _parse_companies_from_xlsx(blob: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(blob), read_only=True, data_only=True)
    except Exception:
        return []

    ws = wb.worksheets[0]

    # Find a header row within the first few rows.
    header_row_idx = None
    header_vals: list[str] = []
    for ridx in range(1, 8):
        vals = []
        ok = False
        for c in ws[ridx]:
            v = str(c.value).strip() if c.value is not None else ""
            vals.append(v)
            if any(k in v for k in ["公司", "企业", "机构", "公司名称", "企业名称"]):
                ok = True
        if ok:
            header_row_idx = ridx
            header_vals = vals
            break

    if header_row_idx is None:
        return []

    def find_col(keys: list[str]) -> int | None:
        for i, h in enumerate(header_vals):
            if not h:
                continue
            if any(k in h for k in keys):
                return i
        return None

    name_col = find_col(["公司", "企业", "机构", "公司名称", "企业名称"])
    url_col = find_col(["网申", "招聘", "入口", "链接", "URL", "url"])

    if name_col is None:
        return []

    rows: list[dict] = []
    for ridx in range(header_row_idx + 1, header_row_idx + 5000):
        row = ws[ridx]
        if not row:
            break
        name_v = row[name_col].value if name_col < len(row) else None
        if name_v is None:
            continue
        name = str(name_v).strip()
        if not name:
            continue
        url = None
        if url_col is not None and url_col < len(row):
            uv = row[url_col].value
            if uv:
                s = str(uv).strip()
                if s.startswith("http://") or s.startswith("https://"):
                    url = s
        rows.append({"name": name[:120], "recruitment_url": url})
    return rows


@router.post("/companies/import", response_class=HTMLResponse)
async def companies_import_post(
    request: Request,
    file: UploadFile = File(...),
    seed_sources: str | None = Form(default=None),
    proxy: str | None = Form(default=None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> HTMLResponse:
    blob = await file.read()
    fname = (file.filename or "").lower()

    rows: list[dict]
    if fname.endswith(".xlsx"):
        rows = _parse_companies_from_xlsx(blob)
    else:
        rows = _parse_companies_from_txt(blob)

    created = 0
    updated = 0
    seeded = 0
    errors: list[str] = []
    touched_company_ids: list[str] = []

    for r in rows[:5000]:
        try:
            c, was_created = _upsert_company(db, r["name"])
            if was_created:
                created += 1
            touched_company_ids.append(c.id)
            # Fill recruitment_url if empty; don't overwrite existing unless it's blank.
            ru = (r.get("recruitment_url") or "").strip()
            if ru and not (c.recruitment_url or "").strip():
                c.recruitment_url = ru
                db.add(c)
                updated += 1
        except Exception as e:
            errors.append(str(e)[:200])

    db.commit()

    if seed_sources:
        # Create Official:* sources for companies that have recruitment_url.
        companies = db.execute(
            select(Company).where(
                Company.id.in_(touched_company_ids),
                Company.recruitment_url.is_not(None),
                Company.recruitment_url != "",
            )
        ).scalars().all()
        for c in companies:
            try:
                src_name = f"Official:{c.name}"
                exists = db.execute(select(CrawlSource).where(CrawlSource.name == src_name)).scalar_one_or_none()
                kind, cfg = _infer_official_source(c.name, c.recruitment_url, proxy=proxy)
                if exists:
                    exists.kind = kind
                    exists.config = cfg
                    exists.enabled = True
                    db.add(exists)
                else:
                    db.add(CrawlSource(kind=kind, name=src_name, enabled=True, config=cfg))
                seeded += 1
            except Exception:
                continue
        db.commit()

    return templates.TemplateResponse(
        "admin_companies_import_result.html",
        {
            "request": request,
            "user": admin,
            "filename": file.filename,
            "parsed": len(rows),
            "created": created,
            "updated": updated,
            "seeded": seeded,
            "errors": errors[:20],
        },
    )


@router.post("/sources/{source_id}/toggle")
def sources_toggle(
    request: Request,
    source_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> RedirectResponse:
    s = db.get(CrawlSource, source_id)
    if s:
        s.enabled = not bool(s.enabled)
        db.add(s)
        db.commit()
    return RedirectResponse(url="/admin/sources", status_code=302)


@router.post("/crawl/run")
def crawl_run_now(
    request: Request,
    since_days: int = Form(default=180),
    mode: str = Form(default="official"),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> HTMLResponse:
    from app.crawler.runner import run

    stats = run(db, since_days=int(since_days), mode=(mode or "core"))
    return templates.TemplateResponse("admin_crawl_result.html", {"request": request, "user": admin, "stats": stats})


@router.post("/sources/{source_id}/run")
def crawl_run_one(
    request: Request,
    source_id: str,
    since_days: int = Form(default=180),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
) -> HTMLResponse:
    from app.crawler.runner import run_one

    stats = run_one(db, source_id=source_id, since_days=int(since_days))
    return templates.TemplateResponse("admin_crawl_result.html", {"request": request, "user": admin, "stats": stats})


